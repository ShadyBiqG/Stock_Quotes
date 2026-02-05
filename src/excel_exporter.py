"""
Экспорт результатов анализа в Excel с форматированием.

Поддерживает два формата:
- Полный формат (export): детальный отчёт с листами Сводка, Детали, Анализ качества
- Упрощённый формат (export_simple): Тикер, Компания, Описание, ответы ИИ по колонкам, итог + график
"""

import logging
from typing import List, Dict
from datetime import date
from pathlib import Path
from collections import Counter
import os
import tempfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Класс для экспорта результатов в Excel"""
    
    # Цветовая схема
    COLORS = {
        'РАСТЕТ': 'C6EFCE',      # Зеленый
        'ПАДАЕТ': 'FFC7CE',      # Красный
        'СТАБИЛЬНА': 'FFEB9C',   # Желтый
        'ОШИБКА': 'FFC7CE',      # Красный
        'ПОДОЗРИТЕЛЬНО': 'FFD9B3', # Оранжевый
        'HEADER': '4472C4'       # Синий
    }
    
    def _calculate_text_lines(self, text: str, column_width: int) -> int:
        """
        Рассчитать количество строк текста с учётом ширины колонки и переносов.
        
        Args:
            text: Текст ячейки
            column_width: Ширина колонки в символах
            
        Returns:
            Количество строк для отображения текста
        """
        if not text:
            return 1
        
        lines = text.split('\n')
        total_lines = 0
        
        for line in lines:
            if not line:
                total_lines += 1
            else:
                # Количество переносов внутри строки (при wrap_text)
                wrapped_lines = max(1, (len(line) + column_width - 1) // column_width)
                total_lines += wrapped_lines
        
        return total_lines
    
    def __init__(self, output_dir: str = "output/exports"):
        """
        Инициализация экспортера
        
        Args:
            output_dir: Директория для сохранения файлов
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Экспортер инициализирован: {self.output_dir}")
    
    def export(self, results: List[Dict], 
               analysis_date: date = None,
               filename: str = None) -> Path:
        """
        Экспорт результатов в Excel
        
        Args:
            results: Результаты анализа
            analysis_date: Дата анализа
            filename: Имя файла (если None - генерируется автоматически)
            
        Returns:
            Путь к созданному файлу
        """
        if analysis_date is None:
            analysis_date = date.today()
        
        if filename is None:
            filename = f"{analysis_date}_analysis.xlsx"
        
        filepath = self.output_dir / filename
        
        logger.info(f"Экспорт результатов в {filepath}")

        return self._export_with_safe_replace(filepath, results)

    def _export_with_safe_replace(self, target_path: Path, results: List[Dict]) -> Path:
        """
        Экспорт в Excel с защитой от блокировки файла на Windows.
        Пишем во временный файл и затем заменяем целевой. Если целевой файл открыт,
        подбираем новое имя вида *_1.xlsx, *_2.xlsx.

        Args:
            target_path: Желаемый путь итогового файла
            results: Результаты анализа

        Returns:
            Фактический путь созданного файла
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)

        attempts = 10
        last_error = None

        for i in range(attempts):
            final_path = target_path if i == 0 else self._with_suffix_counter(target_path, i)

            tmp_path = None
            try:
                fd, tmp_name = tempfile.mkstemp(
                    suffix=final_path.suffix,
                    prefix=f".tmp_{final_path.stem}_",
                    dir=str(final_path.parent)
                )
                os.close(fd)
                tmp_path = Path(tmp_name)

                # Создание Excel файла во временный путь
                with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
                    self._create_summary_sheet(results, writer)
                    self._create_details_sheet(results, writer)
                    self._create_quality_sheet(results, writer)

                # Применение форматирования
                self._apply_formatting(tmp_path)

                # Атомарная замена (или создание, если файла нет)
                os.replace(tmp_path, final_path)

                logger.info(f"Файл создан: {final_path}")
                return final_path

            except PermissionError as e:
                last_error = e
                logger.warning(f"Нет доступа к файлу {final_path} (возможно открыт в Excel). Попытка {i + 1}/{attempts}")

                if tmp_path and tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass

                continue

            except Exception:
                if tmp_path and tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass
                raise

        raise PermissionError(
            f"Не удалось сохранить Excel файл. Проверьте, что файл не открыт и доступна запись: {target_path}"
        ) from last_error

    def _with_suffix_counter(self, path: Path, counter: int) -> Path:
        """
        Построить путь с числовым суффиксом: file.xlsx -> file_1.xlsx

        Args:
            path: Исходный путь
            counter: Номер суффикса

        Returns:
            Новый путь
        """
        return path.with_name(f"{path.stem}_{counter}{path.suffix}")

    def export_simple(self, results: List[Dict], 
                      analysis_date: date = None,
                      filename: str = None,
                      database=None) -> Path:
        """
        Упрощённый экспорт результатов в Excel.
        
        Формат первого листа "Анализ":
        - Тикер, Компания, Описание, Ответы ИИ (каждый в своей колонке), Итог (РАСТЕТ/ПАДАЕТ/СТАБИЛЕН)
        
        Формат второго листа "История цен":
        - График изменения цены за месяц из истории БД
        
        Args:
            results: Результаты анализа
            analysis_date: Дата анализа
            filename: Имя файла (если None - генерируется автоматически)
            database: Объект Database для получения исторических данных
            
        Returns:
            Путь к созданному файлу
        """
        if analysis_date is None:
            analysis_date = date.today()
        
        if filename is None:
            filename = f"{analysis_date}_simple_analysis.xlsx"
        
        filepath = self.output_dir / filename
        
        logger.info(f"Упрощённый экспорт результатов в {filepath}")

        return self._export_simple_with_safe_replace(filepath, results, database)

    def _export_simple_with_safe_replace(self, target_path: Path, results: List[Dict], database=None) -> Path:
        """
        Упрощённый экспорт в Excel с защитой от блокировки файла на Windows.

        Args:
            target_path: Желаемый путь итогового файла
            results: Результаты анализа
            database: Объект Database для получения исторических данных

        Returns:
            Фактический путь созданного файла
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)

        attempts = 10
        last_error = None

        for i in range(attempts):
            final_path = target_path if i == 0 else self._with_suffix_counter(target_path, i)

            tmp_path = None
            try:
                fd, tmp_name = tempfile.mkstemp(
                    suffix=final_path.suffix,
                    prefix=f".tmp_{final_path.stem}_",
                    dir=str(final_path.parent)
                )
                os.close(fd)
                tmp_path = Path(tmp_name)

                # Создание Excel файла во временный путь
                with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
                    self._create_simple_analysis_sheet(results, writer)
                    self._create_price_history_sheet(results, writer, database)

                # Применение форматирования
                self._apply_simple_formatting(tmp_path)

                # Атомарная замена (или создание, если файла нет)
                os.replace(tmp_path, final_path)

                logger.info(f"Файл создан: {final_path}")
                return final_path

            except PermissionError as e:
                last_error = e
                logger.warning(f"Нет доступа к файлу {final_path} (возможно открыт в Excel). Попытка {i + 1}/{attempts}")

                if tmp_path and tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass

                continue

            except Exception:
                if tmp_path and tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass
                raise

        raise PermissionError(
            f"Не удалось сохранить Excel файл. Проверьте, что файл не открыт и доступна запись: {target_path}"
        ) from last_error

    def _create_simple_analysis_sheet(self, results: List[Dict], writer) -> None:
        """
        Создать лист "Анализ" в упрощённом формате.
        
        Колонки: Тикер | Компания | Описание | ИИ1 (полный ответ) | ИИ2 | ИИ3 ... | Итог
        
        Args:
            results: Результаты анализа
            writer: Excel writer
        """
        # Группировка по тикерам
        stocks_data = {}
        all_models = set()
        
        for r in results:
            ticker = r['ticker']
            model_name = r['model_name']
            all_models.add(model_name)
            
            if ticker not in stocks_data:
                stocks_data[ticker] = {
                    'Тикер': ticker,
                    'Компания': r.get('name', ''),
                    'Описание': r.get('description', ''),
                    'models': {},
                    'predictions': {}
                }
            
            # Сохраняем полный текст анализа от модели с ключевыми факторами
            analysis_text = r.get('analysis_text', '')
            key_factors = r.get('key_factors', [])
            
            # Формируем полный текст: анализ + ключевые факторы
            full_text_parts = []
            
            if analysis_text:
                full_text_parts.append(analysis_text)
            
            if key_factors:
                factors_text = '\n'.join([f"• {f}" for f in key_factors])
                full_text_parts.append(f"\nКлючевые факторы:\n{factors_text}")
            
            if full_text_parts:
                full_analysis = '\n'.join(full_text_parts)
            else:
                full_analysis = r.get('prediction', 'Н/Д')
            
            stocks_data[ticker]['models'][model_name] = full_analysis
            stocks_data[ticker]['predictions'][model_name] = r.get('prediction', 'Н/Д')
        
        # Сортируем модели для единообразного порядка колонок
        sorted_models = sorted(all_models)
        
        # Формируем данные для DataFrame
        rows = []
        for ticker, data in stocks_data.items():
            row = {
                'Тикер': data['Тикер'],
                'Компания': data['Компания'],
                'Описание': data['Описание']
            }
            
            # Добавляем полные ответы каждой модели
            predictions = []
            for model in sorted_models:
                analysis = data['models'].get(model, 'Н/Д')
                row[model] = analysis
                
                prediction = data['predictions'].get(model, 'Н/Д')
                if prediction != 'Н/Д':
                    predictions.append(prediction)
            
            # Вычисление итога (консенсус по прогнозам)
            row['Итог'] = self._calculate_simple_consensus(predictions)
            
            rows.append(row)
        
        # Создание DataFrame
        columns = ['Тикер', 'Компания', 'Описание'] + sorted_models + ['Итог']
        df = pd.DataFrame(rows, columns=columns)
        
        # Сохранение в Excel
        df.to_excel(writer, sheet_name='Анализ', index=False)
        
        logger.debug("Создан лист 'Анализ' (упрощённый формат с полными ответами ИИ)")

    def _calculate_simple_consensus(self, predictions: List[str]) -> str:
        """
        Вычисление итогового прогноза на основе ответов моделей.
        
        Args:
            predictions: Список прогнозов от моделей
            
        Returns:
            Итоговый прогноз: РАСТЕТ, ПАДАЕТ или СТАБИЛЕН
        """
        if not predictions:
            return 'Н/Д'
        
        counts = Counter(predictions)
        most_common = counts.most_common(1)[0]
        
        # Если явное большинство
        if most_common[1] > len(predictions) / 2:
            return most_common[0]
        
        # Если равенство голосов - смотрим приоритет
        # РАСТЕТ и ПАДАЕТ важнее чем СТАБИЛЬНА
        if counts.get('РАСТЕТ', 0) > counts.get('ПАДАЕТ', 0):
            return 'РАСТЕТ'
        elif counts.get('ПАДАЕТ', 0) > counts.get('РАСТЕТ', 0):
            return 'ПАДАЕТ'
        
        return 'СТАБИЛЬНА'

    def _create_price_history_sheet(self, results: List[Dict], writer, database=None) -> None:
        """
        Создать лист "История цен" с изменением цены за месяц из БД.
        
        Формат данных для построения графика динамики цен:
        - Колонка "Дата" - ось X для графика
        - Колонки с тикерами - цены для каждой компании (ось Y)
        
        Args:
            results: Результаты анализа
            writer: Excel writer
            database: Объект Database для получения исторических данных
        """
        # Получаем список тикеров
        tickers = set()
        for r in results:
            tickers.add(r['ticker'])
        
        sorted_tickers = sorted(tickers)
        
        # Если нет БД - создаём простую таблицу с текущими данными
        if database is None:
            stocks_data = {}
            for r in results:
                ticker = r['ticker']
                if ticker not in stocks_data:
                    stocks_data[ticker] = r.get('price', 0)
            
            # Формируем одну строку с текущими ценами
            row = {'Дата': date.today().isoformat()}
            for ticker in sorted_tickers:
                row[ticker] = stocks_data.get(ticker, 0)
            
            df = pd.DataFrame([row])
            df.to_excel(writer, sheet_name='История цен', index=False)
            logger.debug("Создан лист 'История цен' (без истории - нет подключения к БД)")
            return
        
        # Получаем исторические данные за месяц для каждого тикера
        # Собираем данные в формате: {дата: {тикер: цена}}
        dates_prices = {}
        
        for ticker in sorted_tickers:
            history = database.get_historical_data(ticker, days=30)
            
            for h in history:
                d = h['date']
                if d not in dates_prices:
                    dates_prices[d] = {}
                
                # Берём только первую цену за день (убираем дубликаты)
                if ticker not in dates_prices[d]:
                    dates_prices[d][ticker] = h['price']
        
        if not dates_prices:
            # Fallback если нет истории - берём текущие данные из results
            stocks_data = {}
            for r in results:
                ticker = r['ticker']
                if ticker not in stocks_data:
                    stocks_data[ticker] = r.get('price', 0)
            
            row = {'Дата': date.today().isoformat()}
            for ticker in sorted_tickers:
                row[ticker] = stocks_data.get(ticker, 0)
            
            df = pd.DataFrame([row])
            df.to_excel(writer, sheet_name='История цен', index=False)
            logger.debug("Создан лист 'История цен' (нет исторических данных)")
            return
        
        # Формируем строки для DataFrame: Дата | Тикер1 | Тикер2 | ...
        rows = []
        for d in sorted(dates_prices.keys()):
            row = {'Дата': d}
            for ticker in sorted_tickers:
                row[ticker] = dates_prices[d].get(ticker, None)
            rows.append(row)
        
        # Создаём DataFrame с колонками: Дата, затем тикеры
        columns = ['Дата'] + sorted_tickers
        df = pd.DataFrame(rows, columns=columns)
        df.to_excel(writer, sheet_name='История цен', index=False)
        
        logger.debug(f"Создан лист 'История цен' ({len(rows)} дат, {len(sorted_tickers)} тикеров)")

    def _apply_simple_formatting(self, filepath: Path) -> None:
        """
        Применить форматирование к упрощённому Excel файлу.
        
        Args:
            filepath: Путь к файлу
        """
        wb = load_workbook(filepath)
        
        # Форматирование листа "Анализ"
        if 'Анализ' in wb.sheetnames:
            self._format_simple_analysis_sheet(wb['Анализ'])
        
        # Форматирование листа "История цен" и добавление графика
        if 'История цен' in wb.sheetnames:
            self._format_price_history_sheet(wb['История цен'])
        
        wb.save(filepath)
        logger.debug("Применено форматирование (упрощённый формат)")

    def _format_simple_analysis_sheet(self, ws) -> None:
        """
        Форматирование листа Анализ (упрощённый формат с полными ответами ИИ).
        
        Args:
            ws: Worksheet объект
        """
        # Определяем индексы колонок
        header_row = [cell.value for cell in ws[1]]
        last_col_idx = len(header_row)  # Индекс колонки "Итог"
        
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'], 
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ширины колонок для расчёта переносов (в символах)
        col_widths = {
            1: 10,   # Тикер
            2: 25,   # Компания
            3: 40,   # Описание
            last_col_idx: 12  # Итог
        }
        # Колонки ИИ по умолчанию 50 символов
        default_ai_width = 50
        
        # Форматирование данных
        for row in ws.iter_rows(min_row=2):
            max_lines = 1  # Минимум 1 строка
            
            for cell in row:
                col_idx = cell.column
                
                # Колонка "Итог" - цветовое кодирование
                if col_idx == last_col_idx and cell.value in self.COLORS:
                    cell.fill = PatternFill(
                        start_color=self.COLORS[cell.value],
                        end_color=self.COLORS[cell.value],
                        fill_type='solid'
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                
                # Тикер - левое выравнивание
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                    cell.font = Font(bold=True)
                
                # Компания, Описание - левое выравнивание с переносом
                elif col_idx <= 3:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Колонки с ответами ИИ - перенос текста
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Расчёт количества строк для динамической высоты
                if cell.value:
                    text = str(cell.value)
                    col_width = col_widths.get(col_idx, default_ai_width)
                    cell_lines = self._calculate_text_lines(text, col_width)
                    max_lines = max(max_lines, cell_lines)
            
            # Динамическая высота строки: 15 пунктов на строку текста + небольшой отступ
            row_height = max(20, max_lines * 15 + 5)
            ws.row_dimensions[row[0].row].height = row_height
        
        # Определяем количество колонок
        header_row = [cell.value for cell in ws[1]]
        num_cols = len(header_row)
        
        # Установка ширины колонок
        for col_idx, column in enumerate(ws.columns, 1):
            column_letter = get_column_letter(col_idx)
            
            if col_idx == 1:  # Тикер
                ws.column_dimensions[column_letter].width = 10
            elif col_idx == 2:  # Компания
                ws.column_dimensions[column_letter].width = 25
            elif col_idx == 3:  # Описание
                ws.column_dimensions[column_letter].width = 40
            elif col_idx == num_cols:  # Итог
                ws.column_dimensions[column_letter].width = 12
            else:  # Колонки с ответами ИИ
                ws.column_dimensions[column_letter].width = 50
        
        # Фиксация заголовка
        ws.freeze_panes = 'A2'

    def _format_price_history_sheet(self, ws) -> None:
        """
        Форматирование листа "История цен" и добавление графика динамики цен по тикерам.
        
        Структура данных на листе:
        - Колонка A: Дата
        - Колонки B, C, D...: Цены тикеров (каждый тикер в своей колонке)
        
        График показывает динамику цены за месяц для каждой компании (тикера).
        
        Args:
            ws: Worksheet объект
        """
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'],
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Форматирование данных
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Форматирование чисел (цены) - 2 знака после запятой
                if cell.column > 1 and cell.value is not None:
                    try:
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        pass
        
        # Установка ширины колонок
        header_row = [cell.value for cell in ws[1]]
        num_cols = len(header_row)
        
        ws.column_dimensions['A'].width = 12  # Дата
        
        for col_idx in range(2, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 12  # Колонки с тикерами
        
        # Создание линейного графика динамики цен по тикерам
        if ws.max_row > 1 and num_cols > 1:
            chart = LineChart()
            chart.title = "Динамика цен за месяц"
            chart.style = 10
            chart.y_axis.title = "Цена"
            chart.x_axis.title = "Дата"
            
            # Данные для графика - все колонки с тикерами (начиная с колонки 2)
            # Каждый тикер будет отдельной линией на графике
            data = Reference(ws, min_col=2, max_col=num_cols, min_row=1, max_row=ws.max_row)
            
            # Категории - даты (колонка A)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Настройка легенды (названия тикеров)
            chart.legend.position = 'b'  # Легенда внизу
            
            # Размер графика
            chart.width = 20
            chart.height = 12
            
            # Размещение графика справа от данных
            chart_position = get_column_letter(num_cols + 2) + "2"
            ws.add_chart(chart, chart_position)
        
        # Фиксация заголовка
        ws.freeze_panes = 'A2'

    def _format_chart_sheet(self, ws) -> None:
        """
        Форматирование листа График и добавление графика накопленной стоимости.
        (Используется в полном формате экспорта)
        
        Args:
            ws: Worksheet объект
        """
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'],
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создание графика накопленной стоимости
        if ws.max_row > 1:
            chart = LineChart()
            chart.title = "Накопленная стоимость портфеля"
            chart.style = 10
            chart.y_axis.title = "Стоимость ($)"
            chart.x_axis.title = "Компании"
            
            # Данные для графика (колонка F - Накопленная стоимость)
            data = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            chart.width = 20
            chart.height = 12
            
            # Размещение графика справа от данных
            ws.add_chart(chart, "H2")
    
    def _create_summary_sheet(self, results: List[Dict], writer) -> None:
        """
        Создать лист "Сводка"
        
        Args:
            results: Результаты анализа
            writer: Excel writer
        """
        # Группировка по тикерам
        summary_data = {}
        
        for r in results:
            ticker = r['ticker']
            
            if ticker not in summary_data:
                summary_data[ticker] = {
                    'Тикер': ticker,
                    'Компания': r.get('name', ''),
                    'Описание': r.get('description', ''),
                    'Сектор': r.get('sector', ''),
                    'Цена': r['price'],
                    'Изм.%': r['change'],
                    'Объем': r['volume']
                }
            
            # Прогнозы моделей
            model_name = r['model_name']
            summary_data[ticker][model_name] = r['prediction']
        
        # Создание DataFrame
        df = pd.DataFrame(list(summary_data.values()))
        
        # Вычисление консенсуса
        model_columns = [col for col in df.columns if col not in [
            'Тикер', 'Компания', 'Описание', 'Сектор', 'Цена', 'Изм.%', 'Объем'
        ]]
        
        def calculate_consensus(row):
            predictions = [row[col] for col in model_columns if pd.notna(row[col])]
            if not predictions:
                return 'Н/Д'
            
            # Если все согласны
            if len(set(predictions)) == 1:
                return predictions[0]
            
            # Большинство
            from collections import Counter
            counts = Counter(predictions)
            most_common = counts.most_common(1)[0]
            
            if most_common[1] > len(predictions) / 2:
                return f"{most_common[0]} ({most_common[1]}/{len(predictions)})"
            
            return f"Разногласие ({most_common[1]}/{len(predictions)})"
        
        df['Консенсус'] = df.apply(calculate_consensus, axis=1)
        
        # Сохранение в Excel
        df.to_excel(writer, sheet_name='Сводка', index=False)
        
        logger.debug("Создан лист 'Сводка'")
    
    def _create_details_sheet(self, results: List[Dict], writer) -> None:
        """
        Создать лист "Детали"
        
        Args:
            results: Результаты анализа
            writer: Excel writer
        """
        details_data = []
        
        for r in results:
            # Полный текст анализа
            analysis_text = r.get('analysis_text', '')
            
            # Ключевые факторы
            key_factors = r.get('key_factors', [])
            if key_factors:
                factors_text = '\n'.join([
                    f"{i+1}. {factor}" 
                    for i, factor in enumerate(key_factors)
                ])
            else:
                # Fallback на старое поле reasons для обратной совместимости
                reasons = r.get('reasons', [])
                factors_text = '\n'.join([
                    f"{i+1}. {reason}" 
                    for i, reason in enumerate(reasons)
                ]) if reasons else 'Не указаны'
            
            details_data.append({
                'Тикер': r['ticker'],
                'Компания': r.get('name', ''),
                'Цена': r['price'],
                'Изм.%': r['change'],
                'Модель': r['model_name'],
                'Прогноз': r['prediction'],
                'Анализ': analysis_text if analysis_text else 'Не предоставлен',
                'Ключевые факторы': factors_text,
                'Уверенность': r['confidence'],
                'Токенов': r.get('tokens_used', 0)
            })
        
        df = pd.DataFrame(details_data)
        df.to_excel(writer, sheet_name='Детали', index=False)
        
        logger.debug("Создан лист 'Детали'")
    
    def _create_quality_sheet(self, results: List[Dict], writer) -> None:
        """
        Создать лист "Анализ качества"
        
        Args:
            results: Результаты анализа
            writer: Excel writer
        """
        # Группировка по тикерам
        stocks = {}
        for r in results:
            ticker = r['ticker']
            if ticker not in stocks:
                stocks[ticker] = []
            stocks[ticker].append(r)
        
        quality_data = []
        
        for ticker, stock_results in stocks.items():
            predictions = [r['prediction'] for r in stock_results]
            
            # Консенсус
            has_consensus = len(set(predictions)) == 1
            
            # Подозрительные ответы
            suspicious_count = sum(
                1 for r in stock_results
                if r.get('validation_flags', {}).get('trust_level') == 'LOW'
            )
            
            # Средняя уверенность
            confidences = [r['confidence'] for r in stock_results]
            conf_map = {'ВЫСОКАЯ': 3, 'СРЕДНЯЯ': 2, 'НИЗКАЯ': 1}
            
            if confidences:
                avg_conf = sum(conf_map.get(c, 1) for c in confidences) / len(confidences)
                avg_conf_text = 'ВЫСОКАЯ' if avg_conf >= 2.5 else (
                    'СРЕДНЯЯ' if avg_conf >= 1.5 else 'НИЗКАЯ'
                )
            else:
                avg_conf_text = 'Н/Д'
            
            quality_data.append({
                'Тикер': ticker,
                'Консенсус': 'Да' if has_consensus else 'Нет',
                'Разных мнений': len(set(predictions)),
                'Подозрительных': suspicious_count,
                'Средняя уверенность': avg_conf_text,
                'Всего токенов': sum(r.get('tokens_used', 0) for r in stock_results)
            })
        
        # Проверка наличия данных
        if not quality_data:
            # Если нет данных, создаем пустой лист с сообщением
            df = pd.DataFrame([{
                'Тикер': 'Нет данных',
                'Консенсус': '-',
                'Разных мнений': '-',
                'Подозрительных': '-',
                'Средняя уверенность': '-',
                'Всего токенов': '-'
            }])
            df.to_excel(writer, sheet_name='Анализ качества', index=False)
            logger.warning("Нет данных для анализа качества")
            return
        
        df = pd.DataFrame(quality_data)
        
        # Итоговая статистика (только если есть данные)
        total_consensus = sum(1 for d in quality_data if d['Консенсус'] == 'Да')
        total_count = len(quality_data)
        avg_opinions = sum(d['Разных мнений'] for d in quality_data) / total_count
        total_suspicious = sum(d['Подозрительных'] for d in quality_data)
        total_tokens = sum(d['Всего токенов'] for d in quality_data)
        
        total_stats = pd.DataFrame([{
            'Тикер': 'ИТОГО',
            'Консенсус': f"{total_consensus} / {total_count}",
            'Разных мнений': f"{avg_opinions:.1f}",
            'Подозрительных': total_suspicious,
            'Средняя уверенность': '-',
            'Всего токенов': total_tokens
        }])
        
        df = pd.concat([df, total_stats], ignore_index=True)
        df.to_excel(writer, sheet_name='Анализ качества', index=False)
        
        logger.debug("Создан лист 'Анализ качества'")
    
    def _apply_formatting(self, filepath: Path) -> None:
        """
        Применить форматирование к Excel файлу
        
        Args:
            filepath: Путь к файлу
        """
        wb = load_workbook(filepath)
        
        # Форматирование листа "Сводка"
        if 'Сводка' in wb.sheetnames:
            self._format_summary_sheet(wb['Сводка'])
        
        # Форматирование листа "Детали"
        if 'Детали' in wb.sheetnames:
            self._format_details_sheet(wb['Детали'])
        
        # Форматирование листа "Анализ качества"
        if 'Анализ качества' in wb.sheetnames:
            self._format_quality_sheet(wb['Анализ качества'])
        
        wb.save(filepath)
        logger.debug("Применено форматирование")
    
    def _format_summary_sheet(self, ws) -> None:
        """Форматирование листа Сводка"""
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'], 
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Цветовое кодирование прогнозов
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value in self.COLORS:
                    cell.fill = PatternFill(
                        start_color=self.COLORS[cell.value],
                        end_color=self.COLORS[cell.value],
                        fill_type='solid'
                    )
                
                # Выравнивание
                if cell.column <= 4:  # Текстовые колонки
                    cell.alignment = Alignment(horizontal='left', 
                                              vertical='top',
                                              wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center',
                                              vertical='center')
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Фиксация заголовка
        ws.freeze_panes = 'A2'
    
    def _format_details_sheet(self, ws) -> None:
        """Форматирование листа Детали"""
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'],
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Найти индексы колонок для текстовых полей
        header_row = [cell.value for cell in ws[1]]
        analysis_col = header_row.index('Анализ') + 1 if 'Анализ' in header_row else None
        factors_col = header_row.index('Ключевые факторы') + 1 if 'Ключевые факторы' in header_row else None
        
        # Перенос текста в анализе и факторах
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                # Колонка "Анализ"
                if analysis_col and cell.column == analysis_col:
                    cell.alignment = Alignment(wrap_text=True,
                                              vertical='top',
                                              horizontal='left')
                    ws.row_dimensions[cell.row].height = 80
                
                # Колонка "Ключевые факторы"
                elif factors_col and cell.column == factors_col:
                    cell.alignment = Alignment(wrap_text=True,
                                              vertical='top',
                                              horizontal='left')
                    ws.row_dimensions[cell.row].height = 80
                
                # Остальные колонки
                else:
                    cell.alignment = Alignment(vertical='center',
                                              horizontal='center')
        
        # Установка ширины колонок
        for i, column in enumerate(ws.columns, 1):
            column_letter = get_column_letter(i)
            
            # Широкие колонки для текстовых полей
            if i == analysis_col:
                ws.column_dimensions[column_letter].width = 70
            elif i == factors_col:
                ws.column_dimensions[column_letter].width = 50
            else:
                # Автоширина для остальных колонок
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Фиксация заголовка
        ws.freeze_panes = 'A2'
    
    def _format_quality_sheet(self, ws) -> None:
        """Форматирование листа Анализ качества"""
        # Заголовки
        header_fill = PatternFill(start_color=self.COLORS['HEADER'],
                                  end_color=self.COLORS['HEADER'],
                                  fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Выделение итоговой строки
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E7E6E6',
                                   end_color='E7E6E6',
                                   fill_type='solid')
        
        # Автоширина
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width


# Пример использования
if __name__ == "__main__":
    from .database import Database
    
    logging.basicConfig(level=logging.INFO)
    
    # Получение результатов из БД
    with Database() as db:
        results = db.get_analysis_results()
    
    # Экспорт
    exporter = ExcelExporter()
    filepath = exporter.export(results)
    
    print(f"✅ Файл создан: {filepath}")
