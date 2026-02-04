"""
Полный цикл: Анализ → БД → Экспорт Excel
"""

import asyncio
import sys
from datetime import date

# Исправление кодировки для Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

from src.database import Database
from src.llm_manager import OpenRouterClient
from src.company_info import CompanyInfoProvider
from src.analyzer import StockAnalyzer
from src.excel_exporter import ExcelExporter
from pathlib import Path
import yaml

def load_config():
    """Загрузка конфигурации"""
    config_dir = Path("config")
    
    # API ключи
    with open(config_dir / "api_keys.yaml", 'r', encoding='utf-8') as f:
        api_keys = yaml.safe_load(f)
    
    # LLM конфигурация
    with open(config_dir / "llm_config.yaml", 'r', encoding='utf-8') as f:
        llm_config = yaml.safe_load(f)
    
    config = llm_config
    config['openrouter'] = {
        'api_key': api_keys.get('openrouter_api_key', ''),
        'base_url': 'https://openrouter.ai/api/v1'
    }
    
    return config


async def main():
    """Главная функция"""
    print("\n" + "="*80)
    print("🔄 ПОЛНЫЙ ЦИКЛ ТЕСТА: АНАЛИЗ → БД → EXCEL")
    print("="*80)
    
    # Загрузка конфигурации
    config = load_config()
    
    # Инициализация компонентов
    print("\n📦 Инициализация...")
    
    llm_client = OpenRouterClient(
        api_key=config['openrouter']['api_key'],
        base_url=config['openrouter']['base_url']
    )
    
    db = Database(config['database']['path'])
    
    company_provider = CompanyInfoProvider(
        cache_duration_days=config['company_info']['cache_duration_days'],
        fallback_llm_client=llm_client if config['company_info']['fallback_to_llm'] else None
    )
    
    analyzer = StockAnalyzer(
        llm_client=llm_client,
        database=db,
        company_provider=company_provider,
        config=config
    )
    
    # Тестовая акция
    test_stock = {
        'ticker': 'TSLA',
        'price': 245.67,
        'change': 3.21,
        'volume': 95_430_000,
        'additional_info': 'Tesla Inc. - производитель электромобилей'
    }
    
    print(f"\n📊 Тестовая акция: {test_stock['ticker']}")
    print(f"   Цена: ${test_stock['price']}")
    print(f"   Изменение: {test_stock['change']:+.2f}%")
    
    # ШАГ 1: Анализ
    print(f"\n🚀 ШАГ 1: Запуск анализа...")
    
    results = await analyzer.analyze_stocks(
        stocks=[test_stock],
        analysis_date=date.today()
    )
    
    print(f"✅ Анализ завершен: {results['successful']} успешно, {results['failed']} ошибок")
    
    # ШАГ 2: Проверка БД
    print(f"\n💾 ШАГ 2: Проверка данных в БД...")
    
    db_results = db.get_analysis_results(
        analysis_date=date.today(),
        ticker=test_stock['ticker']
    )
    
    print(f"✅ Найдено записей в БД: {len(db_results)}")
    
    for result in db_results:
        analysis_text = result.get('analysis_text', '')
        key_factors = result.get('key_factors', [])
        
        print(f"\n   🤖 {result['model_name']}:")
        print(f"      Прогноз: {result['prediction']}")
        print(f"      Анализ: {'✅' if analysis_text else '❌'} ({len(analysis_text)} символов)")
        print(f"      Факторы: {'✅' if key_factors else '❌'} ({len(key_factors)} шт.)")
    
    # ШАГ 3: Экспорт в Excel
    print(f"\n📄 ШАГ 3: Экспорт в Excel...")
    
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"test_{timestamp}.xlsx"
    
    exporter = ExcelExporter()
    filepath = exporter.export(db_results, analysis_date=date.today(), filename=filename)
    
    print(f"✅ Файл создан: {filepath}")
    
    # ШАГ 4: Проверка Excel файла
    print(f"\n🔍 ШАГ 4: Проверка содержимого Excel...")
    
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    
    if 'Детали' in wb.sheetnames:
        ws = wb['Детали']
        header_row = [cell.value for cell in ws[1]]
        
        print(f"   Колонки: {', '.join(header_row)}")
        
        # Проверка первой строки данных
        if ws.max_row > 1:
            row2 = list(ws[2])
            
            analysis_col = header_row.index('Анализ') if 'Анализ' in header_row else None
            factors_col = header_row.index('Ключевые факторы') if 'Ключевые факторы' in header_row else None
            
            print(f"\n   Проверка первой строки данных:")
            print(f"   Тикер: {ws['A2'].value}")
            print(f"   Модель: {ws['E2'].value}")
            print(f"   Прогноз: {ws['F2'].value}")
            
            if analysis_col:
                analysis_value = row2[analysis_col].value
                print(f"   Анализ: {'✅ Заполнен' if analysis_value else '❌ Пусто'}")
                if analysis_value:
                    print(f"      Начало: {str(analysis_value)[:100]}...")
            
            if factors_col:
                factors_value = row2[factors_col].value
                print(f"   Факторы: {'✅ Заполнены' if factors_value else '❌ Пусто'}")
                if factors_value:
                    lines = str(factors_value).split('\n')
                    print(f"      Количество: {len(lines)}")
                    print(f"      Первый: {lines[0][:80]}...")
    
    wb.close()
    
    # Открытие файла
    print(f"\n📂 Открытие файла...")
    import subprocess
    subprocess.Popen(['start', str(filepath)], shell=True)
    
    print("\n" + "="*80)
    print("✅ ПОЛНЫЙ ЦИКЛ ЗАВЕРШЕН УСПЕШНО")
    print("="*80)
    print(f"\n💡 Проверьте лист 'Детали' в файле Excel:")
    print(f"   - Колонка 'Анализ' должна содержать развернутый текст анализа")
    print(f"   - Колонка 'Ключевые факторы' должна содержать 3 пункта")
    
    db.close()


if __name__ == "__main__":
    asyncio.run(main())
