"""
Тестирование экспорта в Excel
"""

import sys
from pathlib import Path
from datetime import date

# Исправление кодировки для Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

from src.database import Database
from src.excel_exporter import ExcelExporter

print("\n" + "="*80)
print("🧪 ТЕСТИРОВАНИЕ ЭКСПОРТА В EXCEL")
print("="*80)

# Подключение к БД
db = Database('data/stocks.db')
print(f"\n✅ Подключено к БД: data/stocks.db")

# Получение результатов
results = db.get_analysis_results(analysis_date=date.today())
print(f"✅ Загружено записей: {len(results)}")

if not results:
    print("\n⚠️  Нет данных для экспорта!")
    print("   Сначала запустите анализ: python main.py")
    sys.exit(1)

# Проверка структуры данных
print(f"\n📊 Проверка структуры данных:")
sample = results[0]
print(f"   Тикер: {sample.get('ticker')}")
print(f"   Модель: {sample.get('model_name')}")
print(f"   Прогноз: {sample.get('prediction')}")
print(f"   Уверенность: {sample.get('confidence')}")

# Проверка новых полей
analysis_text = sample.get('analysis_text', '')
key_factors = sample.get('key_factors', [])

print(f"\n   Анализ: {'✅ Есть' if analysis_text else '❌ Отсутствует'} ({len(analysis_text)} символов)")
print(f"   Факторы: {'✅ Есть' if key_factors else '❌ Отсутствуют'} ({len(key_factors)} шт.)")

if analysis_text:
    print(f"      Начало: {analysis_text[:80]}...")

if key_factors:
    print(f"      Первый фактор: {key_factors[0][:80]}...")

# Экспорт
print(f"\n📄 Экспорт в Excel...")
exporter = ExcelExporter()
filepath = exporter.export(results, analysis_date=date.today())

print(f"\n✅ Файл создан: {filepath}")
print(f"   Размер: {filepath.stat().st_size / 1024:.1f} KB")

# Проверка структуры файла
from openpyxl import load_workbook

wb = load_workbook(filepath)
print(f"\n📋 Листы в файле:")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"   • {sheet_name}: {ws.max_row-1} строк данных")
    
    # Для листа "Детали" показываем колонки
    if sheet_name == 'Детали':
        header_row = [cell.value for cell in ws[1]]
        print(f"     Колонки: {', '.join(header_row)}")
        
        # Проверка заполненности
        if ws.max_row > 1:
            # Проверяем вторую строку
            row2 = list(ws[2])
            analysis_col = header_row.index('Анализ') if 'Анализ' in header_row else None
            factors_col = header_row.index('Ключевые факторы') if 'Ключевые факторы' in header_row else None
            
            if analysis_col is not None:
                analysis_value = row2[analysis_col].value
                print(f"     Анализ заполнен: {'✅ Да' if analysis_value else '❌ Нет'}")
                if analysis_value:
                    print(f"       Начало: {str(analysis_value)[:60]}...")
            
            if factors_col is not None:
                factors_value = row2[factors_col].value
                print(f"     Факторы заполнены: {'✅ Да' if factors_value else '❌ Нет'}")
                if factors_value:
                    print(f"       Начало: {str(factors_value)[:60]}...")

wb.close()

print("\n" + "="*80)
print("✅ ТЕСТ ЗАВЕРШЕН")
print("="*80)
print(f"\n💡 Откройте файл: {filepath}")

db.close()
